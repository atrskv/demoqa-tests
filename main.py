from zipfile import ZipFile
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook


def test_read_pdf_file_from_zip(create_and_delete_resources_directory):

    # GIVEN:
    with ZipFile('resources/pdf_example.zip', mode='w') as pdf_archive:
        pdf_archive.write('files_for_zip/pdf_example.pdf')

    # WHEN:
    with ZipFile('resources/pdf_example.zip') as pdf_file:
        pdf_file.extract('files_for_zip/pdf_example.pdf', 'resources')

    reader = PdfReader('files_for_zip/pdf_example.pdf')
    text_from_page2 = reader.pages[1].extract_text()

    # THEN:
    assert 'The end, and just as well' in text_from_page2, 'Wrong text!'


def test_read_csv_file_from_zip(create_and_delete_resources_directory):

    # GIVEN:
    with ZipFile('resources/csv_example.zip', mode='w') as csv_archive:
        csv_archive.write('files_for_zip/csv_example.csv')

    # WHEN:
    with ZipFile('resources/csv_example.zip') as csv_file:
        csv_file.extract('files_for_zip/csv_example.csv', 'resources')

    with open('files_for_zip/csv_example.csv') as csv_file_:
        reader = csv.reader(csv_file_)

        result = []

        for row in reader:
            result.append(row)

    # THEN:
    assert (
                result ==
            [
                ['Username; Identifier;First name;Last name'],
                ['booker12;9012;Rachel;Booker'],
                ['grey07;2070;Laura;Grey'],
                ['johnson81;4081;Craig;Johnson'],
                ['jenkins46;9346;Mary;Jenkins'],
                ['smith79;5079;Jamie;Smith'],
                []
            ]

            ), 'Wrong values!'


def test_read_xlsx_file_from_zip(create_and_delete_resources_directory):

    # GIVEN:
    with ZipFile('resources/xlsx_example.zip', mode='w') as xlsx_archive:
        xlsx_archive.write('files_for_zip/xlsx_example.xlsx')

    # WHEN:
    with ZipFile('resources/xlsx_example.zip') as xlsx_file:
        xlsx_file.extract('files_for_zip/xlsx_example.xlsx', 'resources')

    workbook = load_workbook('files_for_zip/xlsx_example.xlsx')
    sheet = workbook.active

    result = list(sheet.iter_rows(values_only=True))

    # THEN:
    assert (
                result ==
            [
                (0, 'First Name', 'Last Name', 'Gender', 'Country', 'Age', 'Date', 'Id'),
                (1, 'Dulce', 'Abril', 'Female', 'United States', 32, '15/10/2017', 1562),
                (2, 'Mara', 'Hashimoto', 'Female', 'Great Britain', 25, '16/08/2016', 1582),
                (3, 'Philip', 'Gent', 'Male', 'France', 36, '21/05/2015', 2587),
                (4, 'Kathleen', 'Hanner', 'Female', 'United States', 25, '15/10/2017', 3549),
                (5, 'Nereida', 'Magwood', 'Female', 'United States', 58, '16/08/2016', 2468),
                (6, 'Gaston', 'Brumm', 'Male', 'United States', 24, '21/05/2015', 2554),
                (7, 'Etta', 'Hurn', 'Female', 'Great Britain', 56, '15/10/2017', 3598),
                (8, 'Earlean', 'Melgar', 'Female', 'United States', 27, '16/08/2016', 2456),
                (9, 'Vincenza', 'Weiland', 'Female', 'United States', 40, '21/05/2015', 6548)
            ]
    )
