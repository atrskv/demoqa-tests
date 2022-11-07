# READ CSV
# import csv
#
# with open('example_csv.csv') as f:
#     reader = csv.reader(f)
#     for row in range(len(reader.line_num)):
#         if row == 3:
#             print(row)


# READ XLSX
from openpyxl import load_workbook

workbook = load_workbook('file_example_XLSX_10.xlsx')
sheet = workbook.active  # активируем книгу
print(sheet.cell(row=3, column=2).value)

# READ XLS
import xlrd
book = xlrd.open_workbook('file_example_XLS_10.xls')
print(book.nsheets)
print(book.sheet_names())

sheet = book.sheet_by_index(0)
print(sheet.name)

print(sheet.cell_value(rowx=2, colx=3))

for r in range(sheet.nrows):
    print(sheet.row(r))





